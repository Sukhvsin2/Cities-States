from openpyxl import Workbook
import openpyxl
import csv

# storing all of the states into a dict
# where state: id
states = {}

stateCSV = open("statesCSV.csv", "r")
stateCSV.__next__()
for s in stateCSV:
	id = s.split(",")[0]
	state = s.split(",")[1]
	states[state] = id

stateCSV.close()

#reading us cities

path = "UScities&states.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active


end_city = "El Monte Mobile Village\n"
trigger = True
count = 1
cities = []

while(trigger):
	city = sheet_obj.cell(row = count+1, column=1)
	state = sheet_obj.cell(row = count+1, column = 3)
	lat = sheet_obj.cell(row = count+1, column = 4)
	lng = sheet_obj.cell(row = count+1, column = 5)
	# print(count, city.value, state.value)
	count = count+1
	# to hold state_id & city

	if(city.value == None):
		trigger = False
		break

	temp = []
	state_id = states.get(state.value + "\n")
	temp.append(state_id)
	temp.append(city.value)
	temp.append(state.value)
	temp.append(lat.value)
	temp.append(lng.value)
	cities.append(temp)


workbook = Workbook()
sheet = workbook.active


total = []

sheet["A1"] = "id"
sheet["B1"] = "city"
sheet["C1"] = "state"
sheet["D1"] = "lat"
sheet["E1"] = "lng"

count = 2

with open('newCities.csv', 'w', newline='') as file:
	writer = csv.writer(file)
	fields = ["state_id", "city", "state", "lat", "lng"]

	writer.writerow(fields)

	for l in cities:
		sheet["A${count}".format(count=count)] = l[0]
		sheet["B${count}".format(count=count)] = l[1]
		sheet["C${count}".format(count=count)] = l[2]
		sheet["D${count}".format(count=count)] = l[3]
		sheet["E${count}".format(count=count)] = l[4]
		writer.writerow([l[0], l[1], l[2], l[3], l[4]])
		count = count + 1


print("count: ", count)
workbook.save(filename="newCities.xlsx")